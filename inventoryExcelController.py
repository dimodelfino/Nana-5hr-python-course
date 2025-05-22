import openpyxl


inventory_file = openpyxl.load_workbook('inventory.xlsx')
product_list = inventory_file["Sheet1"]

products_per_supplier = {}
total_value_per_supplier = {}
products_under_10 = {}

#calculation number of products per supplier
for product in range(2, product_list.max_row + 1):
    product_number = product_list.cell(product, 1).value
    inventory = product_list.cell(product, 2).value
    price = product_list.cell(product, 3).value
    supplier_name = product_list.cell(product, 4).value
    inventory_price = product_list.cell(product, 5)

    if supplier_name not in products_per_supplier:
        print("Adding new supplier")
        products_per_supplier[supplier_name] = 1
    else:
        products_per_supplier[supplier_name] += 1

    #calculation total value of inventory per supplier

    product_value = inventory * price

    if supplier_name not in total_value_per_supplier:
        total_value_per_supplier[supplier_name] = product_value
    else:
        total_value_per_supplier[supplier_name] += product_value

    #logic with products with inventory < 10
    if inventory < 10:
        products_under_10[product_number] = inventory

    #add value for total inventory price
    inventory_price.value = product_value

inventory_file.save("inventory_with_total_value.xlsx")

print(products_per_supplier)
print(total_value_per_supplier)
print(products_under_10)
